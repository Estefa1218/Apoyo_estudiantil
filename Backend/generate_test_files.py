#!/usr/bin/env python3
"""
Script para generar archivos Excel de prueba para Sistema PAE
Uso: python3 generate_test_files.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import os

def create_styled_workbook(rows):
    """Crear libro con estilos"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir encabezados
    headers = list(rows[0].keys())
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Escribir datos
    for row_num, row_data in enumerate(rows, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = row_data[header]
    
    # Ajustar ancho de columnas
    for col_num, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
    
    return wb

# ============ ARCHIVO 1: Estándar ============
print("📝 Generando test_estudiantes.xlsx...")
data1 = [
    {"Nombre Completo": "Juan García", "Email": "juan.garcia@mail.com", "Días Ausente": 5},
    {"Nombre Completo": "María López", "Email": "maria.lopez@mail.com", "Días Ausente": 8},
    {"Nombre Completo": "Carlos Pérez", "Email": "carlos.perez@mail.com", "Días Ausente": 3},
    {"Nombre Completo": "Ana González", "Email": "ana.gonzalez@mail.com", "Días Ausente": 6},
    {"Nombre Completo": "Roberto Martínez", "Email": "roberto.martinez@mail.com", "Días Ausente": 2},
]
wb1 = create_styled_workbook(data1)
wb1.save("test_estudiantes.xlsx")
print("  ✅ Creado: test_estudiantes.xlsx")

# ============ ARCHIVO 2: Nombre y Apellido Separados ============
print("📝 Generando test_nombres_separados.xlsx...")
data2 = [
    {"Nombre": "Juan", "Apellido": "García", "Email": "juan.garcia@mail.com", "Días Ausente": 5},
    {"Nombre": "María", "Apellido": "López", "Email": "maria.lopez@mail.com", "Días Ausente": 8},
    {"Nombre": "Carlos", "Apellido": "Pérez", "Email": "carlos.perez@mail.com", "Días Ausente": 3},
]
wb2 = create_styled_workbook(data2)
wb2.save("test_nombres_separados.xlsx")
print("  ✅ Creado: test_nombres_separados.xlsx")

# ============ ARCHIVO 3: Con Última Conexión (Fecha) ============
print("📝 Generando test_fecha_conexion.xlsx...")
hoy = datetime.now()
data3 = [
    {"Nombre Completo": "Juan García", "Email": "juan.garcia@mail.com", "Última Conexión": (hoy - timedelta(days=7)).strftime("%Y-%m-%d")},
    {"Nombre Completo": "María López", "Email": "maria.lopez@mail.com", "Última Conexión": (hoy - timedelta(days=12)).strftime("%Y-%m-%d")},
    {"Nombre Completo": "Carlos Pérez", "Email": "carlos.perez@mail.com", "Última Conexión": (hoy - timedelta(days=4)).strftime("%Y-%m-%d")},
]
wb3 = create_styled_workbook(data3)
wb3.save("test_fecha_conexion.xlsx")
print("  ✅ Creado: test_fecha_conexion.xlsx")

# ============ ARCHIVO 4: Variaciones de Nombres de Columna ============
print("📝 Generando test_variaciones.xlsx...")
data4 = [
    {"Full Name": "Juan García", "correo electronico": "juan.garcia@mail.com", "Days Absent": 5},
    {"Full Name": "María López", "correo electronico": "maria.lopez@mail.com", "Days Absent": 8},
    {"Full Name": "Carlos Pérez", "correo electronico": "carlos.perez@mail.com", "Days Absent": 3},
]
wb4 = create_styled_workbook(data4)
wb4.save("test_variaciones.xlsx")
print("  ✅ Creado: test_variaciones.xlsx")

# ============ ARCHIVO 5: Con Errores (para pruebas de validación) ============
print("📝 Generando test_errores.xlsx...")
data5 = [
    {"Nombre Completo": "Juan García", "Email": "juan.garcia@mail.com", "Días Ausente": 5},
    {"Nombre Completo": "María López", "Email": "", "Días Ausente": 8},  # Sin email
    {"Nombre Completo": "Carlos Pérez", "Email": "invalido.email", "Días Ausente": 3},  # Email sin @
    {"Nombre Completo": "", "Email": "ana@mail.com", "Días Ausente": 6},  # Sin nombre
]
wb5 = create_styled_workbook(data5)
wb5.save("test_errores.xlsx")
print("  ✅ Creado: test_errores.xlsx")

print("")
print("════════════════════════════════════════════════════════")
print("✅ Todos los archivos de prueba han sido creados")
print("════════════════════════════════════════════════════════")
print("")
print("Archivos generados:")
print("  1. test_estudiantes.xlsx - Archivo estándar")
print("  2. test_nombres_separados.xlsx - Con nombre y apellido separados")
print("  3. test_fecha_conexion.xlsx - Con fecha de última conexión")
print("  4. test_variaciones.xlsx - Con variaciones de nombres de columna")
print("  5. test_errores.xlsx - Con errores de validación")
print("")
print("Pruebas rápidas:")
print("  curl -X POST -F \"file=@test_estudiantes.xlsx\" http://localhost:3001/api/upload")
print("")
